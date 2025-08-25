# -*- coding: utf-8 -*-
"""
Avito Price Analyzer — настольное приложение (Windows/macOS/Linux)
Автор: ChatGPT (для пользователя)

Что делает:
1) Загружаете Excel-файл с товарами:
   - Колонка A: бренд
   - Колонка B: название/модель + характеристики (например: "iPhone 12, 128 ГБ, белый")
   - Колонка C: закупочная цена (в рублях)
   Каждая строка — отдельный товар.

2) Нажимаете "Старт". Программа для каждой строки ищет первые 20 объявлений на Avito,
   фильтрует именно ваш товар по ключевым словам и считает среднюю цену.
   В Excel заполняются колонки:
   - D: средняя цена по Avito (₽)
   - E: наценка в % = ((D - C) / C) * 100
   - F: ссылка на самое дешёвое из учтённых объявлений

3) Подсветка строк по E:
   - 5–10%     → жёлтый
   - ≥ 10%     → зелёный
   - < 5%      → без подсветки

⚠️ Важно: сайт Avito регулярно меняет вёрстку и методы защиты от ботов.
Я сделал максимально бережный парсинг через обычные HTTP-запросы с mobile User‑Agent.
Если вдруг Avito изменится и парсер перестанет находить цены — приложение не упадёт, а аккуратно пропустит строки и запишет лог.
"""

import threading
import queue
import time
import random
import re
import sys
import os
import math
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple

import requests
from bs4 import BeautifulSoup

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.workbook import Workbook

import tkinter as tk
from tkinter import ttk, filedialog, messagebox


APP_TITLE = "Avito Price Analyzer"
RESULT_SUFFIX = "_analyzed"
MAX_RESULTS_PER_QUERY = 20
REQUEST_TIMEOUT = 20
REQUEST_DELAY_RANGE = (1.0, 2.0)  # задержка между запросами, чтобы беречь сайт


MOBILE_UAS = [
    # Несколько мобильных User-Agent строк на случай ротации
    "Mozilla/5.0 (Linux; Android 13; Pixel 6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Mobile Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_4 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (Linux; Android 12; SM-G991B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Mobile Safari/537.36",
]

STOPWORDS_RU = {
    "и", "или", "а", "для", "с", "на", "в", "из", "к", "по", "без", "до", "от", "что",
    "гб", "тб", "gb", "tb", "цвет", "белый", "черный", "чёрный", "серый", "серебристый", "красный",
    "новый", "б/у", "бу", "есть", "нет", "про", "про-", "встроенный", "версия"
}

CAPACITY_UNITS = {
    "гб": "gb", "gb": "gb", "тб": "tb", "tb": "tb"
}


@dataclass
class Listing:
    title: str
    url: str
    price_rub: Optional[int]


def log_print(widget: tk.Text, text: str):
    widget.insert(tk.END, text + "\\n")
    widget.see(tk.END)
    widget.update_idletasks()


def normalize_text(s: str) -> str:
    s = s.lower()
    # стандартные замены: гб/тб → gb/tb; чёрный → черный
    s = s.replace("чёр", "чер")
    s = re.sub(r"(\\d+)\\s*(гб|gb)", lambda m: f"{m.group(1)}gb", s)
    s = re.sub(r"(\\d+)\\s*(тб|tb)", lambda m: f"{m.group(1)}tb", s)
    # удаляем лишнее
    s = re.sub(r"[^a-z0-9а-яё\\s\\-]+", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\\s+", " ", s).strip()
    return s


def extract_tokens(brand: str, name: str) -> List[str]:
    full = f"{brand} {name}"
    full = normalize_text(full)
    raw_tokens = re.findall(r"[a-z0-9а-яё\\-]+", full, flags=re.IGNORECASE)
    tokens = []
    for t in raw_tokens:
        if t in STOPWORDS_RU:
            continue
        # игнорируем короткие односимвольные
        if len(t) <= 1:
            continue
        tokens.append(t)
    # Убираем дубликаты, сохраняя порядок
    deduped = []
    seen = set()
    for t in tokens:
        if t not in seen:
            seen.add(t)
            deduped.append(t)
    return deduped


def build_search_url(query: str) -> str:
    # Мобильная версия Avito с поиском по всей России
    # Пример: https://m.avito.ru/rossiya?q=iphone 12 128gb
    from urllib.parse import quote_plus
    return f"https://m.avito.ru/rossiya?q={quote_plus(query)}"


def fetch_html(url: str, session: requests.Session) -> Optional[str]:
    headers = {
        "User-Agent": random.choice(MOBILE_UAS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.7,en;q=0.6",
        "Connection": "keep-alive",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "DNT": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Upgrade-Insecure-Requests": "1",
    }
    try:
        resp = session.get(url, headers=headers, timeout=REQUEST_TIMEOUT)
        if resp.status_code == 200 and "<html" in resp.text.lower():
            return resp.text
        return None
    except requests.RequestException:
        return None


def parse_listings_from_html(html: str, base_url: str) -> List[Listing]:
    """
    Пытаемся быть устойчивыми к изменениям вёрстки: пробуем несколько наборов селекторов.
    """
    soup = BeautifulSoup(html, "lxml")

    # кандидатные контейнеры объявлений
    candidates = []

    # 1) По data-marker (часто используется на Avito)
    candidates.extend(soup.select('[data-marker="item"]'))

    # 2) По семантическим блокам (fallback)
    candidates.extend(soup.select('div[class*="item"]'))
    candidates.extend(soup.select('article'))
    candidates = list(dict.fromkeys(candidates))  # remove duplicates

    listings: List[Listing] = []

    def extract_price_text(node) -> Optional[str]:
        # Несколько вариантов селекторов
        for sel in ['[data-marker="item-price"]', 'span.price-text', 'span[data-marker*="price"]', 'span[itemprop="price"]']:
            el = node.select_one(sel)
            if el and el.get_text(strip=True):
                return el.get_text(" ", strip=True)
        # бэкап: искать по слову "₽"
        text = node.get_text(" ", strip=True)
        if "₽" in text:
            return text
        return None

    def extract_title_and_url(node) -> Tuple[Optional[str], Optional[str]]:
        # Пытаемся найти ссылку и заголовок
        a = node.select_one('a[data-marker="item-title"]') or node.select_one('a[itemprop="url"]') or node.find('a', href=True)
        title = None
        url = None
        if a:
            title = a.get_text(" ", strip=True) or None
            href = a.get("href")
            if href:
                if href.startswith("http"):
                    url = href
                else:
                    # относительная ссылка
                    if href.startswith("/"):
                        url = f"https://m.avito.ru{href}"
                    else:
                        url = base_url.rstrip("/") + "/" + href
        return title, url

    for node in candidates:
        title, url = extract_title_and_url(node)
        if not title or not url:
            continue
        price_text = extract_price_text(node)
        price_val = None
        if price_text:
            digits = re.findall(r"\\d+", price_text.replace("\\xa0", " "))
            if digits:
                price_val = int("".join(digits))
        listings.append(Listing(title=title, url=url, price_rub=price_val))

    # fallback: иногда листинг может быть вне ожидаемых контейнеров
    if not listings:
        for a in soup.find_all("a", href=True):
            t = a.get_text(" ", strip=True)
            if t and len(t) > 3 and "/item" in a.get("href", ""):
                url = a["href"]
                if not url.startswith("http"):
                    url = "https://m.avito.ru" + (url if url.startswith("/") else ("/" + url))
                listings.append(Listing(title=t, url=url, price_rub=None))

    # ограничиваем первыми 50 (потом отфильтруем до 20)
    return listings[:50]


def score_listing_match(listing_title: str, required_tokens: List[str]) -> float:
    """
    Оцениваем, насколько объявление похоже на нужную конфигурацию.
    Простой скоринг: доля токенов, которые встретились в заголовке.
    """
    if not required_tokens:
        return 0.0
    lt = normalize_text(listing_title)
    hits = 0
    for tok in required_tokens:
        if tok in lt:
            hits += 1
    return hits / len(required_tokens)


def choose_relevant(listings: List[Listing], required_tokens: List[str]) -> List[Listing]:
    # Считаем score и отбираем те, где совпадений >= 50% (можно корректировать)
    scored = [(score_listing_match(li.title, required_tokens), li) for li in listings if li.price_rub]
    scored.sort(key=lambda x: x[0], reverse=True)
    filtered = [li for s, li in scored if s >= 0.5]
    # Берём первые 20
    return filtered[:20]


def process_excel(input_path: str, log_widget: tk.Text, progress: ttk.Progressbar) -> Optional[str]:
    try:
        df = pd.read_excel(input_path, header=None)  # без заголовков, ровно A/B/C
    except Exception as e:
        log_print(log_widget, f"Ошибка чтения файла: {e}")
        return None

    # Ожидаем минимум 3 колонки
    if df.shape[1] < 3:
        log_print(log_widget, "Ошибка: в файле должно быть минимум 3 колонки (A,B,C).")
        return None

    # Подготовим D/E/F
    # D: средняя цена, E: наценка %, F: ссылка
    # Приведём закупку к числу
    def to_float(x):
        try:
            if pd.isna(x):
                return math.nan
            s = str(x).replace(" ", "").replace("\\xa0", "").replace(",", ".")
            s = re.sub(r"[^0-9.\\-]", "", s)
            return float(s) if s else math.nan
        except Exception:
            return math.nan

    session = requests.Session()

    total_rows = len(df)
    for idx, row in df.iterrows():
        brand = str(row.iloc[0]) if not pd.isna(row.iloc[0]) else ""
        name = str(row.iloc[1]) if not pd.isna(row.iloc[1]) else ""
        cost_c = to_float(row.iloc[2])

        progress["value"] = int((idx / max(1, total_rows)) * 100)
        progress.update_idletasks()

        if not brand.strip() and not name.strip():
            log_print(log_widget, f"[{idx+1}] Пустая строка — пропуск.")
            df.loc[idx, 3] = None
            df.loc[idx, 4] = None
            df.loc[idx, 5] = None
            continue

        query = f"{brand} {name}".strip()
        tokens = extract_tokens(brand, name)
        log_print(log_widget, f"[{idx+1}] Ищу: {query} | ключи: {', '.join(tokens) if tokens else '—'}")

        url = build_search_url(query)
        html = fetch_html(url, session)

        if not html:
            log_print(log_widget, f"   ⚠️ Не удалось получить результаты (возможно, защита сайта).")
            df.loc[idx, 3] = None
            df.loc[idx, 4] = None
            df.loc[idx, 5] = None
            # небольшая задержка и продолжим
            time.sleep(random.uniform(*REQUEST_DELAY_RANGE))
            continue

        listings_all = parse_listings_from_html(html, base_url="https://m.avito.ru")

        if not listings_all:
            log_print(log_widget, f"   ⚠️ Ничего не найдено по запросу.")
            df.loc[idx, 3] = None
            df.loc[idx, 4] = None
            df.loc[idx, 5] = None
            time.sleep(random.uniform(*REQUEST_DELAY_RANGE))
            continue

        listings = choose_relevant(listings_all, tokens)

        if not listings:
            log_print(log_widget, f"   ⚠️ Нашёл объявления, но ни одно не прошло фильтр по конфигурации.")
            df.loc[idx, 3] = None
            df.loc[idx, 4] = None
            df.loc[idx, 5] = None
            time.sleep(random.uniform(*REQUEST_DELAY_RANGE))
            continue

        prices = [li.price_rub for li in listings if li.price_rub is not None]
        avg_price = round(sum(prices) / len(prices), 2) if prices else None
        cheapest = min(listings, key=lambda li: li.price_rub or 10**12)

        df.loc[idx, 3] = avg_price
        if avg_price is not None and not (cost_c is None or (isinstance(cost_c, float) and math.isnan(cost_c))):
            # Наценка (а не коэффициент!). Так логичнее под ваши пороги 5–10% и 10%+
            markup_percent = (avg_price - cost_c) / cost_c * 100.0 if cost_c != 0 else None
            df.loc[idx, 4] = markup_percent
        else:
            df.loc[idx, 4] = None
        df.loc[idx, 5] = cheapest.url if cheapest and cheapest.url else None

        log_print(log_widget, f"   ✓ Нашёл {len(listings)} объявл.; средняя: {avg_price if avg_price else '—'} ₽; дёш.: {cheapest.price_rub if cheapest.price_rub else '—'} ₽")
        time.sleep(random.uniform(*REQUEST_DELAY_RANGE))

    # Сохраняем новый файл с форматированием через openpyxl
    out_path = make_output_path(input_path)
    try:
        save_with_formatting(df, out_path)
    except Exception as e:
        log_print(log_widget, f"⚠️ Не удалось применить форматирование, сохраню без него: {e}")
        # резерв: просто сохраняем pandas (без подсветки)
        df.to_excel(out_path, header=False, index=False)

    progress["value"] = 100
    progress.update_idletasks()
    return out_path


def make_output_path(input_path: str) -> str:
    base, ext = os.path.splitext(input_path)
    if ext.lower() not in [".xlsx", ".xls"]:
        ext = ".xlsx"
    return f"{base}{RESULT_SUFFIX}{ext}"


def save_with_formatting(df: pd.DataFrame, out_path: str):
    # Пишем в xlsx без заголовков
    tmp_path = out_path
    df.to_excel(tmp_path, header=False, index=False)

    wb = load_workbook(tmp_path)
    ws = wb.active

    # Назначаем ширину столбцов A..F
    col_widths = [20, 50, 14, 16, 12, 50]
    for i, w in enumerate(col_widths, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = w

    # Шапки нет, но можно для удобства добавить их строкой 1 при желании.
    # Здесь оставляем как в исходнике: данные начинаются с первой строки.

    # Формат чисел: C (закупка) и D (средняя) — денежный, E — проценты
    max_row = ws.max_row
    for row in range(1, max_row + 1):
        c_cell = ws.cell(row=row, column=3)
        d_cell = ws.cell(row=row, column=4)
        e_cell = ws.cell(row=row, column=5)

        # Денежный формат ₽
        d_cell.number_format = '#,##0" ₽"'
        c_cell.number_format = '#,##0" ₽"'

        # Проценты с 2 знаками
        e_cell.number_format = '0.00" %"'

        # Подсветка строки по E
        e_val = e_cell.value
        if isinstance(e_val, (int, float)):
            if 5.0 <= e_val < 10.0:
                fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # мягкий жёлтый
                apply_row_fill(ws, row, fill, max_col=6)
            elif e_val >= 10.0:
                fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # мягкий зелёный
                apply_row_fill(ws, row, fill, max_col=6)

    wb.save(tmp_path)


def apply_row_fill(ws, row_idx: int, fill: PatternFill, max_col: int):
    for col in range(1, max_col + 1):
        ws.cell(row=row_idx, column=col).fill = fill


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("840x660")
        self.resizable(True, True)

        # State
        self.input_path: Optional[str] = None
        self.output_path: Optional[str] = None
        self.worker_thread: Optional[threading.Thread] = None
        self.task_queue: "queue.Queue[callable]" = queue.Queue()

        # UI
        self.create_widgets()

        # Periodic UI updater for thread-safe logging
        self.after(100, self.process_queue)

    def create_widgets(self):
        pad = 10

        # Frame: загрузка файла
        frm_top = ttk.Frame(self, padding=pad)
        frm_top.pack(fill="x")

        self.path_var = tk.StringVar()
        ttk.Label(frm_top, text="Файл Excel (A=бренд, B=модель/характеристики, C=закупка ₽):").pack(anchor="w")
        path_row = ttk.Frame(frm_top)
        path_row.pack(fill="x", pady=(5, 0))
        self.path_entry = ttk.Entry(path_row, textvariable=self.path_var)
        self.path_entry.pack(side="left", fill="x", expand=True)
        ttk.Button(path_row, text="Выбрать файл…", command=self.choose_file).pack(side="left", padx=(6, 0))

        # Frame: кнопки
        frm_btns = ttk.Frame(self, padding=pad)
        frm_btns.pack(fill="x")
        self.btn_start = ttk.Button(frm_btns, text="Старт", command=self.on_start, state="disabled")
        self.btn_start.pack(side="left")
        ttk.Button(frm_btns, text="Закрыть", command=self.on_close).pack(side="right")

        # Прогресс
        frm_prog = ttk.Frame(self, padding=pad)
        frm_prog.pack(fill="x")
        ttk.Label(frm_prog, text="Ход обработки:").pack(anchor="w")
        self.progress = ttk.Progressbar(frm_prog, orient="horizontal", length=400, mode="determinate")
        self.progress.pack(fill="x", pady=(5, 0))

        # Лог
        frm_log = ttk.Frame(self, padding=pad)
        frm_log.pack(fill="both", expand=True)
        ttk.Label(frm_log, text="Лог:").pack(anchor="w")
        self.log = tk.Text(frm_log, height=20, wrap="word")
        self.log.pack(fill="both", expand=True)
        self.log.configure(state="normal")

        # Выходной файл
        frm_out = ttk.Frame(self, padding=pad)
        frm_out.pack(fill="x")
        ttk.Label(frm_out, text="Итоговый файл:").pack(anchor="w")
        out_row = ttk.Frame(frm_out)
        out_row.pack(fill="x", pady=(5, 0))
        self.out_var = tk.StringVar(value="—")
        self.out_entry = ttk.Entry(out_row, textvariable=self.out_var)
        self.out_entry.pack(side="left", fill="x", expand=True)
        ttk.Button(out_row, text="Открыть папку", command=self.open_output_folder).pack(side="left", padx=(6, 0))

        self.update_start_state()

    def choose_file(self):
        path = filedialog.askopenfilename(
            title="Выберите Excel-файл",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if path:
            self.input_path = path
            self.path_var.set(path)
            self.update_start_state()

    def update_start_state(self):
        self.btn_start["state"] = "normal" if self.input_path else "disabled"

    def on_start(self):
        if not self.input_path:
            messagebox.showwarning("Внимание", "Сначала выберите Excel-файл.")
            return
        self.progress["value"] = 0
        self.out_var.set("—")
        self.output_path = None
        self.log.delete("1.0", tk.END)
        self.btn_start["state"] = "disabled"

        # Запускаем в отдельном потоке, чтобы GUI не зависал
        def worker():
            try:
                out = process_excel(self.input_path, self.log, self.progress)
                def after_ok():
                    if out:
                        self.output_path = out
                        self.out_var.set(out)
                        messagebox.showinfo("Готово", "Обработка завершена.")
                    else:
                        messagebox.showwarning("Готово", "Обработка завершена с ошибками. См. лог.")
                self.task_queue.put(after_ok)
            except Exception as e:
                def after_err():
                    messagebox.showerror("Ошибка", f"Непредвиденная ошибка: {e}")
                    self.btn_start["state"] = "normal"
                self.task_queue.put(after_err)
            finally:
                def after_final():
                    self.btn_start["state"] = "normal"
                self.task_queue.put(after_final)

        self.worker_thread = threading.Thread(target=worker, daemon=True)
        self.worker_thread.start()

    def process_queue(self):
        try:
            while True:
                cb = self.task_queue.get_nowait()
                cb()
        except queue.Empty:
            pass
        self.after(100, self.process_queue)

    def open_output_folder(self):
        if not self.output_path or not os.path.exists(self.output_path):
            messagebox.showinfo("Инфо", "Итоговый файл пока не создан.")
            return
        folder = os.path.dirname(self.output_path) or "."
        if sys.platform.startswith("win"):
            os.startfile(folder)
        elif sys.platform == "darwin":
            os.system(f'open "{folder}"')
        else:
            os.system(f'xdg-open "{folder}"')

    def on_close(self):
        self.destroy()


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
